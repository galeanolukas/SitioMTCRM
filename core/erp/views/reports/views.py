from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Sum, Count, F, ExpressionWrapper, FloatField, Q
from django.db.models.functions import TruncDay, TruncMonth, TruncYear, Coalesce
from django.utils import timezone
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime
from django.utils import timezone
from datetime import timedelta
import json
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from core.erp.models import Sale, DetSale, Product, Company, Expense
from core.erp.models_report_changes import ReportChangeLog, ReportConfiguration
from core.user.models import User


class UnifiedReportsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'reports/unified_reports.html'
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener parámetros de filtro
        report_type = self.request.GET.get('report_type', 'sales')
        company_id = self.request.GET.get('company', '')
        start_date = self.request.GET.get('start_date', '')
        end_date = self.request.GET.get('end_date', '')
        payment_method = self.request.GET.get('payment_method', '')
        page = self.request.GET.get('page', 1)
        
        # Empresas para el dropdown
        companies = Company.objects.all()
        context['companies'] = companies
        
        # Establecer empresa por defecto si no se especifica
        if not company_id:
            # Para superusuarios, usar la primera empresa o la del usuario si tiene
            if self.request.user.is_superuser:
                user_company = getattr(self.request.user, 'company_id', None)
                if user_company:
                    company_id = str(user_company)
                else:
                    # Si el superusuario no tiene empresa asignada, usar la primera activa
                    first_company = companies.filter(is_active=True).first()
                    company_id = str(first_company.id) if first_company else ''
            else:
                # Para usuarios no superusuarios, usar su empresa
                user_company = getattr(self.request.user, 'company_id', None)
                company_id = str(user_company) if user_company else ''
        
        # Fechas por defecto (últimos 30 días)
        if not start_date or not end_date:
            end_date_obj = timezone.now()
            start_date_obj = end_date_obj - timedelta(days=30)
            start_date = start_date_obj.strftime('%Y-%m-%d')
            end_date = end_date_obj.strftime('%Y-%m-%d')
        
        context.update({
            'report_type': report_type,
            'company_id': company_id,
            'start_date': start_date,
            'end_date': end_date,
            'payment_method': payment_method,
        })
        
        # Obtener datos según el tipo de reporte
        if report_type == 'sales':
            context['sales_data'] = self.get_sales_data(company_id, start_date, end_date, payment_method, page)
        elif report_type == 'inventory':
            context['inventory_data'] = self.get_inventory_data(company_id, page)
        elif report_type == 'inventory_enhanced':
            context['inventory_enhanced_data'] = self.get_inventory_enhanced_data(company_id, page, self.request.GET)
        elif report_type == 'sales_by_period':
            period_type = self.request.GET.get('period_type', 'daily')
            context['sales_by_period_data'] = self.get_sales_by_period_data(company_id, period_type, start_date, end_date)
        elif report_type == 'product_sales':
            product_id = self.request.GET.get('product_id', '')
            context['product_sales_data'] = self.get_product_sales_data(company_id, product_id, start_date, end_date)
        elif report_type == 'expenses':
            context['expenses_data'] = self.get_expenses_data(company_id, start_date, end_date, page)
        elif report_type == 'profit':
            context['profit_data'] = self.get_profit_data(company_id, start_date, end_date)
        elif report_type == 'top_selling':
            context['top_selling_data'] = self.get_top_selling_data(company_id, start_date, end_date, page)
        
        # Agregar logs de cambios para opción de deshacer
        if report_type in ['inventory_enhanced', 'sales_by_period', 'product_sales']:
            context['recent_changes'] = ReportChangeLog.objects.filter(
                report_type=report_type,
                is_reverted=False
            ).order_by('-created_at')[:10]
        
        return context
    
    def get_sales_data(self, company_id, start_date, end_date, payment_method, page=1):
        from core.erp.choices import payment_method_choices
        
        # Convertir strings a datetime para el rango completo
        start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1, seconds=-1)
        
        # Filtros base
        filters = {
            'date_joined__range': [start_datetime, end_datetime],
        }
        
        if company_id:
            filters['company_id'] = company_id
        
        sales_queryset = Sale.objects.filter(**filters).select_related('cli', 'company').order_by('-date_joined')
        
        # Si se filtra por método de pago, incluir ventas combinadas que contengan ese método
        if payment_method:
            from core.erp.choices import payment_method_choices
            pm_map = dict(payment_method_choices)
            method_name = pm_map.get(payment_method, payment_method)
            
            # Filtrar ventas que tengan el método exacto o que contengan el nombre en métodos combinados
            sales_queryset = sales_queryset.filter(
                Q(payment_method=payment_method) | Q(payment_method__icontains=method_name)
            )
        
        # Paginación - 50 ventas por página
        paginator = Paginator(sales_queryset, 50)
        try:
            sales = paginator.page(page)
        except PageNotAnInteger:
            sales = paginator.page(1)
        except EmptyPage:
            sales = paginator.page(paginator.num_pages)
        
        # Resumen (usando todos los datos, no solo la página actual)
        summary = sales_queryset.aggregate(
            total_sales=Sum('total'),
            total_count=Count('id'),
            avg_ticket=Sum('total') / Count('id')
        )
        
        # Por forma de pago
        payment_breakdown = sales_queryset.values('payment_method').annotate(
            count=Count('id'),
            total=Sum('total')
        ).order_by('-total')
        
        # Ventas diarias
        daily_sales = sales_queryset.annotate(
            date=TruncDay('date_joined')
        ).values('date').annotate(
            total=Sum('total'),
            count=Count('id')
        ).order_by('date')
        
        return {
            'sales': sales,
            'summary': summary,
            'payment_breakdown': payment_breakdown,
            'daily_sales': daily_sales,
            'payment_method_choices': dict(payment_method_choices),
        }
    
    def get_inventory_data(self, company_id, page=1):
        filters = {}
        if company_id:
            filters['company_id'] = company_id
        
        products_queryset = Product.objects.filter(**filters).select_related('cat', 'supplier').order_by('name')
        
        # Paginación - 50 productos por página
        paginator = Paginator(products_queryset, 50)
        try:
            products = paginator.page(page)
        except PageNotAnInteger:
            products = paginator.page(1)
        except EmptyPage:
            products = paginator.page(paginator.num_pages)
        
        # Resumen (usando todos los datos, no solo la página actual)
        summary = products_queryset.aggregate(
            total_products=Count('id'),
            total_stock=Sum('stock'),
            total_value=Sum(F('stock') * F('pvp_final'))
        )
        
        return {
            'products': products,
            'summary': summary,
        }
    
    def get_inventory_enhanced_data(self, company_id, page=1, request_params=None):
        """Reporte mejorado de inventario con filtros y análisis detallado"""
        filters = {}
        if company_id:
            filters['company_id'] = company_id
        
        products_queryset = Product.objects.filter(**filters).select_related('cat', 'supplier')
        
        # Aplicar filtros adicionales
        if request_params:
            stock_filter = request_params.get('stock_filter', '')
            category_filter = request_params.get('category', '')
            supplier_filter = request_params.get('supplier', '')
            search_query = request_params.get('search', '')
            
            if stock_filter == 'low':
                products_queryset = products_queryset.filter(stock__lte=F('min_stock'), stock__gt=0)
            elif stock_filter == 'out':
                products_queryset = products_queryset.filter(stock__lte=0)
            elif stock_filter == 'positive':
                products_queryset = products_queryset.filter(stock__gt=0)
            elif stock_filter == 'critical':
                products_queryset = products_queryset.filter(stock__lte=F('min_stock')/2)
            
            if category_filter:
                products_queryset = products_queryset.filter(cat_id=category_filter)
            
            if supplier_filter:
                products_queryset = products_queryset.filter(supplier_id=supplier_filter)
            
            if search_query:
                products_queryset = products_queryset.filter(
                    Q(name__icontains=search_query) | 
                    Q(code__icontains=search_query)
                )
        
        # Paginación - 50 productos por página
        paginator = Paginator(products_queryset, 50)
        try:
            products = paginator.page(page)
        except PageNotAnInteger:
            products = paginator.page(1)
        except EmptyPage:
            products = paginator.page(paginator.num_pages)
        
        # Datos enriquecidos para cada producto
        products_data = []
        for product in products:
            stock_status = product.get_stock_status()
            stock_value = float(product.stock * product.pvp_final)
            cost_value = float(product.stock * product.cost_price)
            potential_profit = float(product.stock * (product.pvp_final - product.cost_price))
            
            products_data.append({
                'id': product.id,
                'name': product.name,
                'code': product.code or 'N/A',
                'category': product.cat.name if product.cat else 'Sin categoría',
                'supplier': product.supplier.name if product.supplier else 'Sin proveedor',
                'stock': float(product.stock),
                'min_stock': float(product.min_stock),
                'stock_status': stock_status,
                'stock_status_display': product.get_stock_status_display(),
                'stock_value': stock_value,
                'cost_value': cost_value,
                'potential_profit': potential_profit,
                'margin_percentage': float(product.margin_percentage),
                'pvp': float(product.pvp),
                'pvp_final': float(product.pvp_final),
                'cost_price': float(product.cost_price),
                'unit': product.unit,
                'unit_display': product.get_unit_display(),
                'track_stock': product.track_stock,
                'has_low_stock': product.has_low_stock(),
                'is_out_of_stock': product.is_out_of_stock(),
                'last_stock_sync': product.last_stock_sync,
            })
        
        # Resumen completo (usando todos los datos, no solo la página actual)
        summary = products_queryset.aggregate(
            total_products=Count('id'),
            total_stock=Sum('stock'),
            total_value=Sum(F('stock') * F('pvp_final')),
            total_cost=Sum(F('stock') * F('cost_price')),
            total_profit=Sum(F('stock') * (F('pvp_final') - F('cost_price'))),
            avg_margin=Avg('margin_percentage'),
            low_stock_count=Count('id', filter=Q(stock__lte=F('min_stock'), stock__gt=0)),
            out_of_stock_count=Count('id', filter=Q(stock__lte=0)),
            critical_stock_count=Count('id', filter=Q(stock__lte=F('min_stock')/2)),
        )
        
        # Análisis por categoría
        category_breakdown = products_queryset.values('cat__name').annotate(
            count=Count('id'),
            total_stock=Sum('stock'),
            total_value=Sum(F('stock') * F('pvp_final')),
            low_stock_count=Count('id', filter=Q(stock__lte=F('min_stock'), stock__gt=0))
        ).order_by('-total_value')
        
        # Análisis por proveedor
        supplier_breakdown = products_queryset.values('supplier__name').annotate(
            count=Count('id'),
            total_stock=Sum('stock'),
            total_value=Sum(F('stock') * F('pvp_final'))
        ).order_by('-total_value')
        
        return {
            'products': products,
            'products_data': products_data,
            'summary': summary,
            'category_breakdown': list(category_breakdown),
            'supplier_breakdown': list(supplier_breakdown),
            'filters_applied': {
                'stock_filter': request_params.get('stock_filter', '') if request_params else '',
                'category': request_params.get('category', '') if request_params else '',
                'supplier': request_params.get('supplier', '') if request_params else '',
                'search': request_params.get('search', '') if request_params else '',
            }
        }
    
    def get_expenses_data(self, company_id, start_date, end_date, page=1):
        filters = {
            'date__range': [start_date, end_date],
        }
        
        if company_id:
            filters['company_id'] = company_id
        
        expenses_queryset = Expense.objects.filter(**filters).select_related('supplier', 'company').order_by('-date')
        
        # Paginación - 50 gastos por página
        paginator = Paginator(expenses_queryset, 50)
        try:
            expenses = paginator.page(page)
        except PageNotAnInteger:
            expenses = paginator.page(1)
        except EmptyPage:
            expenses = paginator.page(paginator.num_pages)
        
        # Resumen (usando todos los datos, no solo la página actual)
        summary = expenses_queryset.aggregate(
            total_expenses=Sum('amount'),
            total_count=Count('id'),
            avg_expense=Sum('amount') / Count('id')
        )
        
        # Por proveedor
        supplier_breakdown = expenses_queryset.values('supplier__name').annotate(
            count=Count('id'),
            total=Sum('amount')
        ).order_by('-total')
        
        # Gastos diarios
        daily_expenses = expenses_queryset.annotate(
            expense_date=TruncDay('date')
        ).values('expense_date').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('expense_date')
        
        return {
            'expenses': expenses,
            'summary': summary,
            'supplier_breakdown': supplier_breakdown,
            'daily_expenses': daily_expenses,
        }
    
    def get_profit_data(self, company_id, start_date, end_date):
        # Calcular ganancias basadas en ventas vs costos
        # Convertir strings a datetime para el rango completo
        start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1, seconds=-1)
        
        filters = {
            'date_joined__range': [start_datetime, end_datetime],
        }
        
        if company_id:
            filters['company_id'] = company_id
        
        # Ventas
        sales = Sale.objects.filter(**filters)
        total_sales = sales.aggregate(total=Sum('total'))['total'] or 0
        
        # Costos de productos vendidos
        sales_details = DetSale.objects.filter(sale__in=sales)
        total_cost = sales_details.aggregate(
            total=Sum(F('cant') * F('prod__cost_price'))
        )['total'] or 0
        
        # Gastos
        expenses = Expense.objects.filter(
            date__range=[start_date, end_date],
            **({'company_id': company_id} if company_id else {})
        )
        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
        
        # Ganancias
        gross_profit = total_sales - total_cost
        net_profit = gross_profit - total_expenses
        
        return {
            'total_sales': total_sales,
            'total_cost': total_cost,
            'total_expenses': total_expenses,
            'gross_profit': gross_profit,
            'net_profit': net_profit,
            'profit_margin': (net_profit / total_sales * 100) if total_sales > 0 else 0,
        }
    
    def get_top_selling_data(self, company_id, start_date, end_date, page=1):
        """Obtener datos de productos más vendidos"""
        # Convertir strings a datetime para el rango completo
        start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1, seconds=-1)
        
        # Filtros base
        filters = {
            'sale__date_joined__range': [start_datetime, end_datetime],
        }
        
        if company_id:
            filters['sale__company_id'] = company_id
        
        # Consulta principal de productos más vendidos
        queryset = DetSale.objects.filter(**filters)
        
        # Agrupar por producto y calcular totales
        top_products = queryset.values(
            'prod_id',
            'prod__name',
            'prod__code',
            'prod__stock'
        ).annotate(
            total_quantity=Sum('cant'),
            total_sales=Count('sale_id', distinct=True),
            total_amount=Sum(F('cant') * F('price')),
            avg_price=Coalesce(Sum(F('cant') * F('price')) / Sum('cant'), 0, output_field=FloatField())
        ).order_by('-total_quantity')
        
        # Convertir cantidades a números formateados por página
        paginator = Paginator(top_products, 50)
        try:
            products_page = paginator.page(page)
        except PageNotAnInteger:
            products_page = paginator.page(1)
        except EmptyPage:
            products_page = paginator.page(paginator.num_pages)
        
        # Convertir a lista de diccionarios para el template
        products_data = []
        for item in products_page:
            if item['total_quantity'] and item['total_quantity'] > 0:
                products_data.append({
                    'id': item['prod_id'],
                    'name': item['prod__name'] or 'Sin nombre',
                    'code': item['prod__code'] or 'N/A',
                    'stock_available': float(item['prod__stock'] or 0),
                    'total_quantity': float(item['total_quantity']),
                    'total_sales': int(item['total_sales'] or 0),
                    'total_amount': float(item['total_amount'] or 0),
                    'avg_price': float(item['avg_price'])
                })
        
        # Resumen (usando todos los datos, no solo la página actual)
        summary = queryset.aggregate(
            total_products=Count('prod_id', distinct=True),
            total_quantity=Sum('cant'),
            total_sales=Count('sale_id', distinct=True),
            total_amount=Sum(F('cant') * F('price')),
            avg_price=Coalesce(Sum(F('cant') * F('price')) / Sum('cant'), 0, output_field=FloatField())
        )
        
        return {
            'products': products_data,
            'products_json': json.dumps(products_data, ensure_ascii=False),
            'summary': summary,
            'page_obj': products_page,
        }
    
    def get_sales_by_period_data(self, company_id, period_type='daily', start_date=None, end_date=None):
        """Reporte de ventas con agregación por período (diario/semanal/mensual)"""
        # Convertir strings a datetime para el rango completo
        if start_date:
            start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        else:
            start_datetime = timezone.now() - timedelta(days=30)
        
        if end_date:
            end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1, seconds=-1)
        else:
            end_datetime = timezone.now()
        
        # Filtros base
        filters = {
            'date_joined__range': [start_datetime, end_datetime],
        }
        
        if company_id:
            filters['company_id'] = company_id
        
        # Determinar función de truncado según período
        if period_type == 'daily':
            trunc_func = TruncDay('date_joined')
            date_format = '%Y-%m-%d'
        elif period_type == 'weekly':
            trunc_func = TruncDay('date_joined')  # Luego agrupar por semana
            date_format = '%Y-%m-%d'
        elif period_type == 'monthly':
            trunc_func = TruncMonth('date_joined')
            date_format = '%Y-%m'
        else:
            trunc_func = TruncDay('date_joined')
            date_format = '%Y-%m-%d'
        
        # Ventas agregadas por período
        sales_by_period = Sale.objects.filter(**filters).annotate(
            period=trunc_func
        ).values('period').annotate(
            total_sales=Count('id'),
            total_amount=Sum('total'),
            total_items=Sum('detsale__cant'),
            avg_ticket=Coalesce(Avg('total'), 0, output_field=FloatField()),
            subtotal=Sum('subtotal'),
            iva=Sum('iva')
        ).order_by('period')
        
        # Procesar datos según tipo de período
        if period_type == 'weekly':
            processed_data = self._group_by_week(sales_by_period)
        else:
            processed_data = []
            for item in sales_by_period:
                processed_data.append({
                    'period': item['period'].strftime(date_format) if item['period'] else '',
                    'total_sales': item['total_sales'] or 0,
                    'total_amount': float(item['total_amount'] or 0),
                    'total_items': float(item['total_items'] or 0),
                    'avg_ticket': float(item['avg_ticket'] or 0),
                    'subtotal': float(item['subtotal'] or 0),
                    'iva': float(item['iva'] or 0),
                })
        
        # Resumen general
        summary = Sale.objects.filter(**filters).aggregate(
            total_sales=Count('id'),
            total_amount=Sum('total'),
            total_items=Sum('detsale__cant'),
            avg_ticket=Coalesce(Avg('total'), 0, output_field=FloatField()),
            best_day=Max('total'),
            worst_day=Min('total')
        )
        
        # Análisis por método de pago
        payment_breakdown = Sale.objects.filter(**filters).values('payment_method').annotate(
            count=Count('id'),
            amount=Sum('total'),
            percentage=Count('id') * 100.0 / Count('id', filter=Q(company_id=company_id))
        ).order_by('-amount')
        
        # Top productos del período
        top_products = DetSale.objects.filter(
            sale__date_joined__range=[start_datetime, end_datetime]
        )
        if company_id:
            top_products = top_products.filter(sale__company_id=company_id)
        
        top_products = top_products.values(
            'prod__name', 'prod__code'
        ).annotate(
            total_quantity=Sum('cant'),
            total_amount=Sum(F('cant') * F('price')),
            sales_count=Count('sale_id', distinct=True)
        ).order_by('-total_quantity')[:10]
        
        return {
            'period_data': processed_data,
            'summary': summary,
            'payment_breakdown': list(payment_breakdown),
            'top_products': list(top_products),
            'period_type': period_type,
            'date_range': {
                'start': start_datetime.strftime('%Y-%m-%d'),
                'end': end_datetime.strftime('%Y-%m-%d'),
            }
        }
    
    def _group_by_week(self, daily_data):
        """Agrupar datos diarios por semana"""
        weekly_summary = {}
        
        for item in daily_data:
            if not item['period']:
                continue
                
            week_start = self._get_week_start(item['period'])
            week_key = week_start.strftime('%Y-%m-%d')
            
            if week_key not in weekly_summary:
                weekly_summary[week_key] = {
                    'period': f"Semana del {week_start.strftime('%d/%m/%Y')}",
                    'total_sales': 0,
                    'total_amount': 0,
                    'total_items': 0,
                    'avg_ticket': 0,
                    'subtotal': 0,
                    'iva': 0,
                    'days_included': []
                }
            
            week_data = weekly_summary[week_key]
            week_data['total_sales'] += item['total_sales'] or 0
            week_data['total_amount'] += float(item['total_amount'] or 0)
            week_data['total_items'] += float(item['total_items'] or 0)
            week_data['subtotal'] += float(item['subtotal'] or 0)
            week_data['iva'] += float(item['iva'] or 0)
            week_data['days_included'].append(item['period'].strftime('%Y-%m-%d'))
            
            # Calcular promedio del ticket
            if week_data['total_sales'] > 0:
                week_data['avg_ticket'] = week_data['total_amount'] / week_data['total_sales']
        
        return list(weekly_summary.values())
    
    def _get_week_start(self, date):
        """Obtener el inicio de la semana para una fecha dada (lunes)"""
        # Asumir que la semana empieza el lunes
        weekday = date.weekday()
        if weekday == 0:  # Ya es lunes
            return date
        else:
            return date - timedelta(days=weekday)
    
    def get_product_sales_data(self, company_id, product_id=None, start_date=None, end_date=None):
        """Reporte de ventas por producto (general o específico)"""
        # Convertir strings a datetime para el rango completo
        if start_date:
            start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        else:
            start_datetime = timezone.now() - timedelta(days=30)
        
        if end_date:
            end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1, seconds=-1)
        else:
            end_datetime = timezone.now()
        
        # Filtros base
        filters = {
            'sale__date_joined__range': [start_datetime, end_datetime],
        }
        
        if company_id:
            filters['sale__company_id'] = company_id
        
        # Base de datos de detalles de ventas
        queryset = DetSale.objects.filter(**filters).select_related('prod', 'sale')
        
        if product_id:
            # Reporte específico de un producto
            filters['prod_id'] = product_id
            product_queryset = queryset.filter(prod_id=product_id)
            
            # Ventas diarias del producto
            daily_sales = product_queryset.annotate(
                sale_date=TruncDay('sale__date_joined')
            ).values('sale_date').annotate(
                daily_quantity=Sum('cant'),
                daily_amount=Sum(F('cant') * F('price')),
                daily_sales=Count('sale_id', distinct=True),
                daily_avg_price=Coalesce(Sum(F('cant') * F('price')) / Sum('cant'), 0, output_field=FloatField())
            ).order_by('sale_date')
            
            # Información del producto
            try:
                product_info = Product.objects.get(pk=product_id)
            except Product.DoesNotExist:
                return {'error': 'Producto no encontrado'}
            
            # Evolución de precios
            price_evolution = product_queryset.annotate(
                sale_date=TruncDay('sale__date_joined')
            ).values('sale_date', 'price').annotate(
                sales_at_price=Count('id')
            ).order_by('sale_date')
            
            # Resumen completo
            summary = product_queryset.aggregate(
                total_quantity=Sum('cant'),
                total_amount=Sum(F('cant') * F('price')),
                total_sales=Count('sale_id', distinct=True),
                avg_price=Coalesce(Sum(F('cant') * F('price')) / Sum('cant'), 0, output_field=FloatField()),
                max_price=Max('price'),
                min_price=Min('price'),
                first_sale=Min('sale__date_joined'),
                last_sale=Max('sale__date_joined')
            )
            
            return {
                'product_info': {
                    'id': product_info.id,
                    'name': product_info.name,
                    'code': product_info.code or 'N/A',
                    'category': product_info.cat.name if product_info.cat else 'Sin categoría',
                    'current_stock': float(product_info.stock),
                    'min_stock': float(product_info.min_stock),
                    'pvp': float(product_info.pvp),
                    'pvp_final': float(product_info.pvp_final),
                    'cost_price': float(product_info.cost_price),
                    'unit': product_info.unit,
                },
                'daily_sales': [
                    {
                        'date': item['sale_date'].strftime('%Y-%m-%d') if item['sale_date'] else '',
                        'quantity': float(item['daily_quantity'] or 0),
                        'amount': float(item['daily_amount'] or 0),
                        'sales': item['daily_sales'] or 0,
                        'avg_price': float(item['daily_avg_price'] or 0),
                    }
                    for item in daily_sales
                ],
                'price_evolution': list(price_evolution),
                'summary': summary,
                'date_range': {
                    'start': start_datetime.strftime('%Y-%m-%d'),
                    'end': end_datetime.strftime('%Y-%m-%d'),
                }
            }
        else:
            # Reporte general de todos los productos
            product_summary = queryset.values(
                'prod_id',
                'prod__name',
                'prod__code',
                'prod__cat__name'
            ).annotate(
                total_quantity=Sum('cant'),
                total_amount=Sum(F('cant') * F('price')),
                total_sales=Count('sale_id', distinct=True),
                avg_price=Coalesce(Sum(F('cant') * F('price')) / Sum('cant'), 0, output_field=FloatField()),
                max_price=Max('price'),
                min_price=Min('price'),
                first_sale=Min('sale__date_joined'),
                last_sale=Max('sale__date_joined')
            ).order_by('-total_amount')
            
            # Convertir a lista de diccionarios
            products_data = []
            for item in product_summary:
                if item['total_quantity'] and item['total_quantity'] > 0:
                    products_data.append({
                        'id': item['prod_id'],
                        'name': item['prod__name'] or 'Sin nombre',
                        'code': item['prod__code'] or 'N/A',
                        'category': item['prod__cat__name'] or 'Sin categoría',
                        'total_quantity': float(item['total_quantity']),
                        'total_amount': float(item['total_amount'] or 0),
                        'total_sales': item['total_sales'] or 0,
                        'avg_price': float(item['avg_price'] or 0),
                        'max_price': float(item['max_price'] or 0),
                        'min_price': float(item['min_price'] or 0),
                        'first_sale': item['first_sale'].strftime('%Y-%m-%d') if item['first_sale'] else '',
                        'last_sale': item['last_sale'].strftime('%Y-%m-%d') if item['last_sale'] else '',
                    })
            
            # Resumen general
            summary = queryset.aggregate(
                total_products=Count('prod_id', distinct=True),
                grand_total_quantity=Sum('cant'),
                grand_total_amount=Sum(F('cant') * F('price')),
                grand_total_sales=Count('sale_id', distinct=True),
                avg_product_price=Coalesce(Sum(F('cant') * F('price')) / Sum('cant'), 0, output_field=FloatField())
            )
            
            return {
                'products': products_data,
                'summary': summary,
                'date_range': {
                    'start': start_datetime.strftime('%Y-%m-%d'),
                    'end': end_datetime.strftime('%Y-%m-%d'),
                }
            }


class ExportReportView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        # Allow superusers and users with view permissions
        return self.request.user.is_superuser or (
            hasattr(self.request.user, 'has_perm') and 
            (self.request.user.has_perm('erp.view_sale') or 
             self.request.user.has_perm('erp.view_product') or
             self.request.user.has_perm('erp.view_expense'))
        )
    
    def get(self, request):
        report_type = request.GET.get('report_type', 'sales')
        company_id = request.GET.get('company', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        payment_method = request.GET.get('payment_method', '')
        export_format = request.GET.get('format', 'csv')
        
        try:
            # Preparar información del período para incluir en los reportes
            period_info = {
                'start_date': start_date,
                'end_date': end_date,
                'company_id': company_id,
                'payment_method': payment_method
            }
            
            # Obtener datos según el tipo de reporte
            if report_type == 'sales':
                data = self.get_sales_export_data(company_id, start_date, end_date, payment_method)
                filename = f'ventas_{start_date}_al_{end_date}'
            elif report_type == 'inventory':
                data = self.get_inventory_export_data(company_id)
                filename = f'inventario_{timezone.now().strftime("%Y-%m-%d")}'
                period_info['start_date'] = timezone.now().strftime('%Y-%m-%d')
                period_info['end_date'] = timezone.now().strftime('%Y-%m-%d')
            elif report_type == 'expenses':
                data = self.get_expenses_export_data(company_id, start_date, end_date)
                filename = f'gastos_{start_date}_al_{end_date}'
            elif report_type == 'profit':
                data = self.get_profit_export_data(company_id, start_date, end_date)
                filename = f'ganancias_{start_date}_al_{end_date}'
            elif report_type == 'top_selling':
                data = self.get_top_selling_export_data(company_id, start_date, end_date)
                filename = f'productos_mas_vendidos_{start_date}_al_{end_date}'
            
            # Exportar según formato
            if export_format == 'excel':
                return self.export_to_excel(data, filename, report_type, period_info)
            elif export_format == 'pdf':
                return self.export_to_pdf(data, filename, report_type, period_info)
            else:
                return self.export_to_csv(data, filename, report_type, period_info)
        except Exception as e:
            # Return error as plain text for debugging
            return HttpResponse(f"Error en exportación: {str(e)}", content_type='text/plain')
    
    def get_sales_export_data(self, company_id, start_date, end_date, payment_method):
        # Convertir strings a datetime para el rango completo
        start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1, seconds=-1)
        
        filters = {
            'date_joined__range': [start_datetime, end_datetime],
        }
        
        if company_id:
            filters['company_id'] = company_id
        
        sales_queryset = Sale.objects.filter(**filters).select_related('cli', 'company')
        
        # Si se filtra por método de pago, incluir ventas combinadas que contengan ese método
        if payment_method:
            from core.erp.choices import payment_method_choices
            pm_map = dict(payment_method_choices)
            method_name = pm_map.get(payment_method, payment_method)
            
            # Filtrar ventas que tengan el método exacto o que contengan el nombre en métodos combinados
            sales_queryset = sales_queryset.filter(
                Q(payment_method=payment_method) | Q(payment_method__icontains=method_name)
            )
        
        return sales_queryset
    
    def get_inventory_export_data(self, company_id):
        filters = {}
        if company_id:
            filters['company_id'] = company_id
        
        return Product.objects.filter(**filters).select_related('cat', 'supplier')
    
    def get_expenses_export_data(self, company_id, start_date, end_date):
        filters = {
            'date__range': [start_date, end_date],
        }
        
        if company_id:
            filters['company_id'] = company_id
        
        return Expense.objects.filter(**filters).select_related('supplier', 'company')
    
    def get_profit_export_data(self, company_id, start_date, end_date):
        # Similar a get_profit_data pero para exportación
        unified_view = UnifiedReportsView()
        return unified_view.get_profit_data(company_id, start_date, end_date)
    
    def get_top_selling_export_data(self, company_id, start_date, end_date):
        """Obtener datos para exportación de productos más vendidos"""
        unified_view = UnifiedReportsView()
        # Para exportación, necesitamos todos los datos sin paginación
        return unified_view.get_top_selling_data(company_id, start_date, end_date, page=1)
    
    def export_to_csv(self, data, filename, report_type, period_info=None):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Agregar información del período al inicio del reporte
        if period_info:
            writer.writerow(['REPORTE DE VENTAS' if report_type == 'sales' else 
                           'REPORTE DE INVENTARIO' if report_type == 'inventory' else
                           'REPORTE DE GASTOS' if report_type == 'expenses' else
                           'REPORTE DE GANANCIAS' if report_type == 'profit' else
                           'REPORTE DE PRODUCTOS MÁS VENDIDOS'])
            writer.writerow([])  # Fila vacía
            
            # Información del período
            if period_info.get('start_date') and period_info.get('end_date'):
                writer.writerow(['Período:', f'Desde {period_info["start_date"]} hasta {period_info["end_date"]}'])
            
            # Información de la empresa si está disponible
            if period_info.get('company_id'):
                try:
                    company = Company.objects.get(id=period_info['company_id'])
                    writer.writerow(['Empresa:', company.name])
                except Company.DoesNotExist:
                    pass
            
            # Método de pago si está disponible
            if period_info.get('payment_method'):
                writer.writerow(['Método de Pago:', period_info['payment_method']])
            
            writer.writerow(['Fecha de Generación:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow([])  # Fila vacía
        
        if report_type == 'sales':
            # Encabezados de columnas
            writer.writerow(['Fecha', 'Ticket/Factura', 'Cliente', 'Subtotal', 'IVA', 'Total', 'Forma de Pago', 'Empresa'])
            total_amount = 0
            total_iva = 0
            for sale in data:
                ticket_factura = ''
                if sale.invoice_number:
                    ticket_factura = f"{sale.invoice_pos}-{sale.invoice_number}"
                else:
                    ticket_factura = f"Ticket #{sale.id}"
                    
                writer.writerow([
                    sale.date_joined.strftime('%Y-%m-%d %H:%M'),
                    ticket_factura,
                    sale.cli.names if sale.cli else 'N/A',
                    float(sale.subtotal),
                    float(sale.iva),
                    float(sale.total),
                    sale.get_payment_method_display(),
                    sale.company.name if sale.company else 'N/A'
                ])
                total_amount += float(sale.total)
                total_iva += float(sale.iva)
            
            # Resumen final
            writer.writerow([])  # Fila vacía
            writer.writerow(['RESUMEN'])
            writer.writerow(['Total Ventas', total_amount])
            writer.writerow(['Total IVA', total_iva])
            writer.writerow(['Cantidad de Ventas', len(data)])
            writer.writerow(['Promedio por Venta', total_amount / len(data) if data else 0])
            
            # Totales por forma de pago
            writer.writerow([])  # Fila vacía
            writer.writerow(['VENTAS POR FORMA DE PAGO'])
            
            # Calcular totales por forma de pago
            from django.db.models import Sum
            from core.erp.choices import payment_method_choices
            payment_totals = data.values('payment_method').annotate(
                total=Sum('total'),
                count=Count('id')
            ).order_by('-total')
            
            for payment in payment_totals:
                payment_name = dict(payment_method_choices).get(payment['payment_method'], payment['payment_method'])
                writer.writerow([payment_name, f"{payment['total']:.2f}", f"({payment['count']} ventas)"])  
        
        elif report_type == 'inventory':
            # Título del reporte
            writer.writerow(['REPORTE DE INVENTARIO'])
            writer.writerow([])  # Fila vacía
            writer.writerow(['Producto', 'Código', 'Categoría', 'Stock', 'Costo', 'Precio Final', 'Valor Total', 'Valor con IVA'])
            total_stock = 0
            total_value = 0
            total_value_with_iva = 0
            for product in data:
                product_value = float(product.stock * product.pvp_final)
                product_value_with_iva = float(product.stock * product.pvp_final)  # pvp_final ya incluye IVA
                writer.writerow([
                    product.name,
                    product.code or '',
                    product.cat.name if product.cat else '',
                    product.stock,
                    product.cost_price,
                    product.pvp_final,
                    product_value,
                    product_value_with_iva
                ])
                total_stock += float(product.stock)
                total_value += product_value
                total_value_with_iva += product_value_with_iva
            
            # Resumen final
            writer.writerow([])  # Fila vacía
            writer.writerow(['RESUMEN'])
            writer.writerow(['Total Productos', len(data)])
            writer.writerow(['Stock Total', total_stock])
            writer.writerow(['Valor Total del Inventario', total_value])
            writer.writerow(['Valor Total con IVA', total_value_with_iva])
        
        elif report_type == 'expenses':
            # Título del reporte
            writer.writerow(['REPORTE DE GASTOS'])
            writer.writerow([])  # Fila vacía
            writer.writerow(['Fecha', 'Descripción', 'Monto', 'Proveedor', 'Empresa'])
            total_amount = 0
            for expense in data:
                writer.writerow([
                    expense.date,
                    expense.description or '',
                    expense.amount,
                    expense.supplier.name if expense.supplier else 'N/A',
                    expense.company.name if expense.company else 'N/A'
                ])
                total_amount += float(expense.amount)
            
            # Resumen final
            writer.writerow([])  # Fila vacía
            writer.writerow(['RESUMEN'])
            writer.writerow(['Total Gastos', total_amount])
            writer.writerow(['Cantidad de Gastos', len(data)])
            writer.writerow(['Promedio por Gasto', total_amount / len(data) if data else 0])
        
        elif report_type == 'profit':
            # Título del reporte
            writer.writerow(['REPORTE DE GANANCIAS'])
            writer.writerow([])  # Fila vacía
            writer.writerow(['Concepto', 'Monto'])
            writer.writerow(['Ventas Totales', data['total_sales']])
            writer.writerow(['Costo de Ventas', data['total_cost']])
            writer.writerow(['Gastos', data['total_expenses']])
            writer.writerow(['Ganancia Bruta', data['gross_profit']])
            writer.writerow(['Ganancia Neta', data['net_profit']])
            writer.writerow(['Margen de Ganancia %', f"{data['profit_margin']:.2f}%"])
        
        elif report_type == 'top_selling':
            # Encabezados de columnas
            writer.writerow(['Código', 'Producto', 'Stock Disponible', 'Total Vendidos', 'N° Ventas', 'Precio Promedio', 'Total Recaudado'])
            
            for product in data['products']:
                writer.writerow([
                    product['code'],
                    product['name'],
                    product['stock_available'],
                    int(product['total_quantity']),
                    product['total_sales'],
                    round(product['avg_price'], 2),
                    round(product['total_amount'], 2)
                ])
            
            # Resumen final
            writer.writerow([])  # Fila vacía
            writer.writerow(['RESUMEN'])
            writer.writerow(['Total de Productos', data['summary']['total_products']])
            writer.writerow(['Cantidad Total Vendida', int(data['summary']['total_quantity'] or 0)])
            writer.writerow(['Número Total de Ventas', data['summary']['total_sales']])
            writer.writerow(['Monto Total Recaudado', data['summary']['total_amount']])
            writer.writerow(['Precio Promedio General', data['summary']['avg_price']])
        
        return response
    
    def export_to_excel(self, data, filename, report_type, period_info=None):
        if not OPENPYXL_AVAILABLE:
            return HttpResponse("Error: openpyxl no está instalado. Instale con: pip install openpyxl", content_type='text/plain')
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        
        # Estilos
        title_font = Font(bold=True, size=14)
        title_alignment = Alignment(horizontal='center')
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        summary_font = Font(bold=True)
        
        row = 1
        
        # Agregar información del período al inicio del reporte
        if period_info:
            # Título del reporte
            report_title = 'REPORTE DE VENTAS' if report_type == 'sales' else \
                         'REPORTE DE INVENTARIO' if report_type == 'inventory' else \
                         'REPORTE DE GASTOS' if report_type == 'expenses' else \
                         'REPORTE DE GANANCIAS' if report_type == 'profit' else \
                         'REPORTE DE PRODUCTOS MÁS VENDIDOS'
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1, value=report_title)
            ws.cell(row=row, column=1).font = title_font
            ws.cell(row=row, column=1).alignment = title_alignment
            row += 2
            
            # Información del período
            if period_info.get('start_date') and period_info.get('end_date'):
                ws.cell(row=row, column=1, value='Período:')
                ws.cell(row=row, column=2, value=f'Desde {period_info["start_date"]} hasta {period_info["end_date"]}')
                row += 1
            
            # Información de la empresa si está disponible
            if period_info.get('company_id'):
                try:
                    company = Company.objects.get(id=period_info['company_id'])
                    ws.cell(row=row, column=1, value='Empresa:')
                    ws.cell(row=row, column=2, value=company.name)
                    row += 1
                except Company.DoesNotExist:
                    pass
            
            # Método de pago si está disponible
            if period_info.get('payment_method'):
                ws.cell(row=row, column=1, value='Método de Pago:')
                ws.cell(row=row, column=2, value=period_info['payment_method'])
                row += 1
            
            # Fecha de generación
            ws.cell(row=row, column=1, value='Fecha de Generación:')
            ws.cell(row=row, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
            row += 2  # Espacio extra después del encabezado
        
        if report_type == 'sales':
            # Encabezados de columnas
            headers = ['Fecha', 'Ticket/Factura', 'Cliente', 'Subtotal', 'IVA', 'Total', 'Forma de Pago', 'Empresa']
            ws.append(headers)
            
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            total_amount = 0
            total_iva = 0
            for sale in data:
                row += 1
                ticket_factura = ''
                if sale.invoice_number:
                    ticket_factura = f"{sale.invoice_pos}-{sale.invoice_number}"
                else:
                    ticket_factura = f"Ticket #{sale.id}"
                    
                ws.append([
                    sale.date_joined.strftime('%Y-%m-%d %H:%M'),
                    ticket_factura,
                    sale.cli.names if sale.cli else 'N/A',
                    float(sale.subtotal),
                    float(sale.iva),
                    float(sale.total),
                    sale.get_payment_method_display(),
                    sale.company.name if sale.company else 'N/A'
                ])
                total_amount += float(sale.total)
                total_iva += float(sale.iva)
            
            # Resumen final
            row += 2
            ws.append(['RESUMEN'])
            ws.cell(row=row, column=1).font = summary_font
            row += 1
            ws.append(['Total Ventas', total_amount])
            row += 1
            ws.append(['Total IVA', total_iva])
            row += 1
            ws.append(['Cantidad de Ventas', len(data)])
            row += 1
            ws.append(['Promedio por Venta', total_amount / len(data) if data else 0])
            
            # Totales por forma de pago
            row += 2
            ws.append(['VENTAS POR FORMA DE PAGO'])
            ws.cell(row=row, column=1).font = summary_font
            row += 1
            
            # Calcular totales por forma de pago
            from django.db.models import Sum
            payment_totals = data.values('payment_method').annotate(
                total=Sum('total'),
                count=Count('id')
            ).order_by('-total')
            
            for payment in payment_totals:
                from core.erp.choices import payment_method_choices
                payment_name = dict(payment_method_choices).get(payment['payment_method'], payment['payment_method'])
                ws.append([payment_name, float(payment['total']), f"({payment['count']} ventas)"])
                row += 1
        
        elif report_type == 'inventory':
            # Título del reporte
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1, value='REPORTE DE INVENTARIO')
            ws.cell(row=row, column=1).font = title_font
            ws.cell(row=row, column=1).alignment = title_alignment
            row += 2
            
            headers = ['Producto', 'Código', 'Categoría', 'Stock', 'Costo', 'Precio Final', 'Valor Total', 'Valor con IVA']
            ws.append(headers)
            
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            total_stock = 0
            total_value = 0
            total_value_with_iva = 0
            for product in data:
                row += 1
                # Manejar valores nulos para evitar error de conversión a float
                stock_value = float(product.stock) if product.stock is not None else 0.0
                pvp_final_value = float(product.pvp_final) if product.pvp_final is not None else 0.0
                cost_price_value = float(product.cost_price) if product.cost_price is not None else 0.0
                
                product_value = stock_value * pvp_final_value
                product_value_with_iva = stock_value * pvp_final_value  # pvp_final ya incluye IVA
                ws.append([
                    product.name,
                    product.code or '',
                    product.cat.name if product.cat else '',
                    stock_value,
                    cost_price_value,
                    pvp_final_value,
                    product_value,
                    product_value_with_iva
                ])
                total_stock += stock_value
                total_value += product_value
                total_value_with_iva += product_value_with_iva
            
            # Resumen final
            row += 2
            ws.append(['RESUMEN'])
            ws.cell(row=row, column=1).font = summary_font
            row += 1
            ws.append(['Total Productos', len(data)])
            row += 1
            ws.append(['Stock Total', total_stock])
            row += 1
            ws.append(['Valor Total del Inventario', total_value])
            row += 1
            ws.append(['Valor Total con IVA', total_value_with_iva])
        
        elif report_type == 'expenses':
            # Título del reporte
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value='REPORTE DE GASTOS')
            ws.cell(row=row, column=1).font = title_font
            ws.cell(row=row, column=1).alignment = title_alignment
            row += 2
            
            headers = ['Fecha', 'Descripción', 'Monto', 'Proveedor', 'Empresa']
            ws.append(headers)
            
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            total_amount = 0
            for expense in data:
                row += 1
                ws.append([
                    expense.date.strftime('%Y-%m-%d'),
                    expense.description or '',
                    float(expense.amount),
                    expense.supplier.name if expense.supplier else 'N/A',
                    expense.company.name if expense.company else 'N/A'
                ])
                total_amount += float(expense.amount)
            
            # Resumen final
            row += 2
            ws.append(['RESUMEN'])
            ws.cell(row=row, column=1).font = summary_font
            row += 1
            ws.append(['Total Gastos', total_amount])
            row += 1
            ws.append(['Cantidad de Gastos', len(data)])
            row += 1
            ws.append(['Promedio por Gasto', total_amount / len(data) if data else 0])
        
        elif report_type == 'profit':
            # Título del reporte
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=1, value='REPORTE DE GANANCIAS')
            ws.cell(row=row, column=1).font = title_font
            ws.cell(row=row, column=1).alignment = title_alignment
            row += 2
            
            headers = ['Concepto', 'Monto']
            ws.append(headers)
            
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            profit_data = [
                ['Ventas Totales', float(data['total_sales'])],
                ['Costo de Ventas', float(data['total_cost'])],
                ['Gastos', float(data['total_expenses'])],
                ['Ganancia Bruta', float(data['gross_profit'])],
                ['Ganancia Neta', float(data['net_profit'])],
                ['Margen de Ganancia %', f"{data['profit_margin']:.2f}%"]
            ]
            
            for item in profit_data:
                row += 1
                ws.append(item)
        
        elif report_type == 'top_selling':
            # Encabezados de columnas
            headers = ['Código', 'Producto', 'Stock Disponible', 'Total Vendidos', 'N° Ventas', 'Precio Promedio', 'Total Recaudado']
            ws.append(headers)
            
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            for product in data['products']:
                row += 1
                ws.append([
                    product['code'],
                    product['name'],
                    product['stock_available'],
                    int(product['total_quantity']),
                    product['total_sales'],
                    round(product['avg_price'], 2),
                    round(product['total_amount'], 2)
                ])
            
            # Resumen final
            row += 2
            ws.append(['RESUMEN'])
            ws.cell(row=row, column=1).font = summary_font
            row += 1
            ws.append(['Total de Productos', data['summary']['total_products']])
            row += 1
            ws.append(['Cantidad Total Vendida', int(data['summary']['total_quantity'] or 0)])
            row += 1
            ws.append(['Número Total de Ventas', data['summary']['total_sales']])
            row += 1
            ws.append(['Monto Total Recaudado', data['summary']['total_amount']])
            row += 1
            ws.append(['Precio Promedio General', data['summary']['avg_price']])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            # Obtener la primera celda no fusionada para obtener el column_letter
            column_letter = None
            for cell in column:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    column_letter = cell.column_letter
                    break
            
            if column_letter:
                for cell in column:
                    try:
                        if not isinstance(cell, openpyxl.cell.cell.MergedCell) and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response.write(excel_file.read())
        return response
    
    def export_to_pdf(self, data, filename, report_type, period_info=None):
        """Exportar datos a PDF"""
        if not REPORTLAB_AVAILABLE:
            return HttpResponse("Error: ReportLab no está instalado. Instale con: pip install reportlab", content_type='text/plain')
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        # Crear documento PDF
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        story = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        # Estilo para encabezados
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        # Estilo para información del período
        info_style = ParagraphStyle(
            'CustomInfo',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            leftIndent=20
        )
        
        # Agregar información del período al inicio del reporte
        if period_info:
            # Título del reporte
            report_title = 'REPORTE DE VENTAS' if report_type == 'sales' else \
                         'REPORTE DE INVENTARIO' if report_type == 'inventory' else \
                         'REPORTE DE GASTOS' if report_type == 'expenses' else \
                         'REPORTE DE GANANCIAS' if report_type == 'profit' else \
                         'REPORTE DE PRODUCTOS MÁS VENDIDOS'
            
            story.append(Paragraph(report_title, title_style))
            
            # Información del período
            if period_info.get('start_date') and period_info.get('end_date'):
                story.append(Paragraph(f"<b>Período:</b> Desde {period_info['start_date']} hasta {period_info['end_date']}", info_style))
            
            # Información de la empresa si está disponible
            if period_info.get('company_id'):
                try:
                    company = Company.objects.get(id=period_info['company_id'])
                    story.append(Paragraph(f"<b>Empresa:</b> {company.name}", info_style))
                except Company.DoesNotExist:
                    pass
            
            # Método de pago si está disponible
            if period_info.get('payment_method'):
                story.append(Paragraph(f"<b>Método de Pago:</b> {period_info['payment_method']}", info_style))
            
            # Fecha de generación
            story.append(Paragraph(f"<b>Fecha de Generación:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", info_style))
            story.append(Spacer(1, 20))  # Espacio después del encabezado
        
        if report_type == 'sales':
            # Datos de la tabla de ventas
            table_data = [['ID', 'Fecha', 'Cliente', 'Total', 'Método de Pago', 'Estado']]
            
            for sale in data:
                table_data.append([
                    str(sale.id),
                    sale.date_joined.strftime('%d/%m/%Y %H:%M') if sale.date_joined else '',
                    sale.cli.name if sale.cli else 'N/A',
                    f"${sale.total:,.2f}",
                    dict(sale.payment_method_choices).get(sale.payment_method, sale.payment_method),
                    'Facturada' if sale.is_invoiced else 'Pendiente'
                ])
            
            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Resumen de ventas
            story.append(Paragraph("RESUMEN DE VENTAS", header_style))
            story.append(Spacer(1, 12))
            
            total_sales = data.count()
            total_amount = data.aggregate(total=Sum('total'))['total'] or 0
            
            summary_data = [
                ['Concepto', 'Valor'],
                ['Total de Ventas', str(total_sales)],
                ['Monto Total', f"${total_amount:,.2f}"],
                ['Promedio por Venta', f"${total_amount/total_sales if total_sales > 0 else 0:,.2f}"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            
        elif report_type == 'inventory':
            # Datos de la tabla de inventario
            table_data = [['Código', 'Producto', 'Categoría', 'Stock', 'Precio Venta', 'Proveedor']]
            
            for product in data:
                table_data.append([
                    product.code or '',
                    product.name,
                    product.cat.name if product.cat else 'N/A',
                    f"{product.stock or 0:.2f}",
                    f"${product.pvp_final or product.pvp or 0:,.2f}",
                    product.supplier.name if product.supplier else 'N/A'
                ])
            
            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Resumen de inventario
            story.append(Paragraph("RESUMEN DE INVENTARIO", header_style))
            story.append(Spacer(1, 12))
            
            total_products = data.count()
            total_stock = data.aggregate(total=Sum('stock'))['total'] or 0
            total_value = data.aggregate(
                total=Sum(F('stock') * F('pvp_final'))
            )['total'] or 0
            
            summary_data = [
                ['Concepto', 'Valor'],
                ['Total de Productos', str(total_products)],
                ['Stock Total', f"{total_stock:.2f}"],
                ['Valor Total del Inventario', f"${total_value:,.2f}"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            
        elif report_type == 'expenses':
            # Datos de la tabla de gastos
            table_data = [['Fecha', 'Proveedor', 'Descripción', 'Monto', 'Pagado por']]
            
            for expense in data:
                table_data.append([
                    expense.date.strftime('%d/%m/%Y') if expense.date else '',
                    expense.supplier.name if expense.supplier else 'N/A',
                    expense.description or '',
                    f"${expense.amount:,.2f}",
                    expense.payer or 'N/A'
                ])
            
            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Resumen de gastos
            story.append(Paragraph("RESUMEN DE GASTOS", header_style))
            story.append(Spacer(1, 12))
            
            total_expenses = data.count()
            total_amount = data.aggregate(total=Sum('amount'))['total'] or 0
            
            summary_data = [
                ['Concepto', 'Valor'],
                ['Total de Gastos', str(total_expenses)],
                ['Monto Total', f"${total_amount:,.2f}"],
                ['Promedio por Gasto', f"${total_amount/total_expenses if total_expenses > 0 else 0:,.2f}"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            
        elif report_type == 'profit':
            # Datos de ganancias (suponiendo que data es un diccionario con información)
            if isinstance(data, dict) and 'sales_data' in data:
                table_data = [['Concepto', 'Ventas', 'Costos', 'Ganancias', 'Margen %']]
                
                for item in data.get('sales_data', []):
                    table_data.append([
                        item.get('product', 'N/A'),
                        f"${item.get('sales', 0):,.2f}",
                        f"${item.get('cost', 0):,.2f}",
                        f"${item.get('profit', 0):,.2f}",
                        f"{item.get('margin', 0):.2f}%"
                    ])
            else:
                # Resumen general de ganancias
                table_data = [['Período', 'Ventas Totales', 'Costos Totales', 'Ganancias', 'Margen %']]
                
                # Calcular totales (esto es un ejemplo, ajustar según la estructura real de datos)
                total_sales = data.aggregate(total=Sum('total'))['total'] or 0 if hasattr(data, 'aggregate') else 0
                total_costs = 0  # Ajustar según el cálculo real de costos
                total_profit = total_sales - total_costs
                margin_percent = (total_profit / total_sales * 100) if total_sales > 0 else 0
                
                table_data.append([
                    f"{period_info.get('start_date', '')} al {period_info.get('end_date', '')}",
                    f"${total_sales:,.2f}",
                    f"${total_costs:,.2f}",
                    f"${total_profit:,.2f}",
                    f"{margin_percent:.2f}%"
                ])
            
            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
        elif report_type == 'top_selling':
            # Datos de la tabla (código existente)
            table_data = [['Código', 'Producto', 'Stock Disponible', 'Total Vendidos', 'Precio Promedio', 'Total Recaudado']]
            
            for product in data['products']:
                table_data.append([
                    product['code'],
                    product['name'],
                    f"{product['stock_available']:.2f}",
                    f"{product['total_quantity']:.2f}",
                    f"${product['avg_price']:.2f}",
                    f"${product['total_amount']:.2f}"
                ])
            
            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Resumen
            story.append(Paragraph("RESUMEN", header_style))
            story.append(Spacer(1, 12))
            
            summary_data = [
                ['Concepto', 'Valor'],
                ['Total de Productos', str(data['summary']['total_products'])],
                ['Cantidad Total Vendida', f"{data['summary']['total_quantity']:.2f}"],
                ['Monto Total Recaudado', f"${data['summary']['total_amount']:.2f}"],
                ['Precio Promedio General', f"${data['summary']['avg_price']:.2f}"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
        
        # Construir PDF
        doc.build(story)
        return response