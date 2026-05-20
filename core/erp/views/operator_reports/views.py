from django.shortcuts import render
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from core.erp.mixins import ValidatePermissionRequiredMixin
from django.db.models import Sum, Count, F
from django.utils import timezone
from datetime import datetime, timedelta
from core.erp.models import Sale, Company, DetSale
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import json

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class OperatorSalesReportView(LoginRequiredMixin, ValidatePermissionRequiredMixin, TemplateView):
    template_name = 'operator_reports/sales_report.html'
    permission_required = 'erp.view_sale'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Reporte de Ventas'
        context['companies'] = Company.objects.all()
        
        # Get company from session or user
        active_cid = self.request.session.get('company_id')
        if not self.request.user.is_superuser:
            active_cid = active_cid or getattr(self.request.user, 'company_id', None)
        context['active_company_id'] = active_cid
        
        return context

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')
            
            if action == 'get_report':
                report_type = request.POST.get('report_type', 'daily')
                report_format = request.POST.get('report_format', 'payment')
                company_id = request.POST.get('company', '')
                start_date = request.POST.get('start_date', '')
                end_date = request.POST.get('end_date', '')
                
                # Get company filter
                active_cid = self.request.session.get('company_id')
                if not self.request.user.is_superuser:
                    active_cid = active_cid or getattr(self.request.user, 'company_id', None)
                
                if company_id and (self.request.user.is_superuser or not active_cid):
                    active_cid = company_id if company_id else active_cid
                
                # Get dates
                if report_type == 'daily':
                    start_date = timezone.now().date().strftime('%Y-%m-%d')
                    end_date = start_date
                elif report_type == 'weekly':
                    end_date = timezone.now().date()
                    start_date = (end_date - timedelta(days=7)).strftime('%Y-%m-%d')
                elif report_type == 'monthly':
                    end_date = timezone.now().date()
                    start_date = (end_date - timedelta(days=30)).strftime('%Y-%m-%d')
                
                # Build filters
                filters = {}
                if active_cid:
                    filters['company_id'] = active_cid
                if start_date and end_date:
                    filters['date_joined__date__range'] = [start_date, end_date]
                
                # Get sales data
                sales = Sale.objects.filter(**filters).order_by('-date_joined')
                
                # Si el formato es ranking de productos, calcular agrupación por producto
                if report_format == 'products':
                    from django.db.models import Avg
                    
                    # Obtener detalles de venta agrupados por producto
                    try:
                        from core.erp.models import Product
                        
                        product_sales = DetSale.objects.filter(
                            sale__in=sales
                        ).values('prod__name', 'prod__id').annotate(
                            total_quantity=Sum('cant'),
                            avg_price=Avg('price'),
                            total_sales=Sum('subtotal')
                        ).order_by('-total_quantity')
                        
                        product_sales_data = []
                        for idx, ps in enumerate(product_sales, 1):
                            # Obtener stock actual del producto
                            try:
                                product = Product.objects.get(id=ps['prod__id'])
                                stock_actual = float(product.stock or 0)
                                # Debug: mostrar valores
                                print(f"Producto: {ps['prod__name']}, ID: {ps['prod__id']}, Stock: {stock_actual}")
                            except Product.DoesNotExist:
                                stock_actual = 0
                                print(f"Producto no encontrado: ID {ps['prod__id']}")
                            except Exception as e:
                                stock_actual = 0
                                print(f"Error obteniendo stock para {ps['prod__name']}: {e}")
                            
                            product_sales_data.append({
                                'rank': idx,
                                'product': ps['prod__name'],
                                'quantity': int(float(ps['total_quantity'])),  # Convertir a entero
                                'price': float(ps['avg_price']),
                                'total': float(ps['total_sales']),
                                'stock': int(stock_actual)  # Stock como entero
                            })
                        
                        data = {
                            'success': True,
                            'report_format': 'products',
                            'products': product_sales_data,
                            'period_type': report_type,
                            'start_date': start_date,
                            'end_date': end_date
                        }
                    except Exception as e:
                        print(f"Error en consulta de productos: {e}")
                        data = {
                            'success': False,
                            'error': f'Error al procesar productos: {str(e)}'
                        }
                else:
                    # Calculate totals (formato original por método de pago)
                    total_sales = sales.aggregate(
                        total=Sum('total'),
                        subtotal=Sum('subtotal'),
                        iva=Sum('iva'),
                        count=Count('id')
                    )
                
                    # Calculate totals by payment method
                    payment_totals = {}
                    payment_choices = {
                        'cash': 'Efectivo',
                        'card': 'Tarjeta',
                        'transfer': 'Transferencia',
                        'mp': 'Mercado Pago',
                        'check': 'Cheque',
                        'combined': 'Combinada'
                    }
                    
                    for method_key, method_name in payment_choices.items():
                        method_total = sales.filter(payment_method=method_key).aggregate(
                            total=Sum('total'),
                            subtotal=Sum('subtotal'),
                            iva=Sum('iva'),
                            count=Count('id')
                        )
                        payment_totals[method_key] = {
                            'name': method_name,
                            'total': float(method_total['total'] or 0),
                            'subtotal': float(method_total['subtotal'] or 0),
                            'iva': float(method_total['iva'] or 0),
                            'count': method_total['count'] or 0
                        }
                    
                    # Prepare data for response
                    sales_data = []
                    for sale in sales:
                        # Get invoice/ticket number
                        if sale.is_invoiced:
                            ticket_number = sale.invoice_number
                        else:
                            # Usar siempre el ID local para mantener consistencia
                            if hasattr(sale, 'local_sale_id') and sale.local_sale_id:
                                ticket_number = f"TK-{sale.local_sale_id:06d}"
                            else:
                                ticket_number = f"TK-{sale.id:06d}"
                        
                        sales_data.append({
                            'id': sale.id,
                            'date': sale.date_joined.strftime('%d/%m/%Y %H:%M'),
                            'client': sale.cli.names if sale.cli else 'Anónimo',
                            'subtotal': float(sale.subtotal),  # PVP puro sin IVA
                            'payment_method': sale.payment_method,  # Enviar el código, no el display
                            'payment_details': getattr(sale, 'payment_details', []),  # Enviar detalles si existen
                            'company': sale.company.name if sale.company else 'N/A',
                            'ticket_number': ticket_number
                        })
                    
                    # Calcular totales generales usando subtotal (PVP puro)
                    grand_total = sum(float(sale.subtotal) for sale in sales)
                    
                    data = {
                        'success': True,
                        'report_format': 'payment',
                        'sales': sales_data,
                        'total_amount': grand_total,
                        'total_count': total_sales['count'] or 0,
                        'payment_totals': payment_totals,
                        'period_type': report_type,
                        'start_date': start_date,
                        'end_date': end_date
                    }
                
            else:
                data['error'] = 'Acción no válida'
                
        except Exception as e:
            data['error'] = str(e)
            
        return JsonResponse(data, safe=False)


@login_required
def operator_sales_export(request):
    """Export sales data for operators"""
    if not request.user.has_perm('erp.view_sale'):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    
    try:
        report_type = request.GET.get('report_type', 'daily')
        report_format = request.GET.get('report_format', 'payment')
        company_id = request.GET.get('company', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        export_format = request.GET.get('format', 'csv')
        
        # Get company filter
        active_cid = request.session.get('company_id')
        if not request.user.is_superuser:
            active_cid = active_cid or getattr(request.user, 'company_id', None)
        
        if company_id and (request.user.is_superuser or not active_cid):
            active_cid = company_id if company_id else active_cid
        
        # Get dates
        if report_type == 'daily':
            start_date = timezone.now().date().strftime('%Y-%m-%d')
            end_date = start_date
        elif report_type == 'weekly':
            end_date = timezone.now().date()
            start_date = (end_date - timedelta(days=7)).strftime('%Y-%m-%d')
        elif report_type == 'monthly':
            end_date = timezone.now().date()
            start_date = (end_date - timedelta(days=30)).strftime('%Y-%m-%d')
        
        # Build filters
        filters = {}
        if active_cid:
            filters['company_id'] = active_cid
        if start_date and end_date:
            filters['date_joined__date__range'] = [start_date, end_date]
        
        # Get sales data
        sales = Sale.objects.filter(**filters).order_by('-date_joined')
        
        if export_format == 'pdf':
            from django.http import HttpResponse
            if not REPORTLAB_AVAILABLE:
                return HttpResponse("Error: ReportLab no está instalado. Instale con: pip install reportlab", content_type='text/plain')
            
            if report_format == 'products':
                return generate_products_pdf_report(sales, start_date, end_date, active_cid, request.user, report_type)
            else:
                return generate_pdf_report(sales, start_date, end_date, active_cid, request.user, report_type)
        
        elif export_format == 'csv':
            import csv
            from django.http import HttpResponse
            
            # Si es formato de productos, generar CSV de ranking
            if report_format == 'products':
                from django.db.models import Avg
                
                product_sales = DetSale.objects.filter(
                    sale__in=sales
                ).values('prod__name', 'prod__id').annotate(
                    total_quantity=Sum('cant'),
                    avg_price=Avg('price'),
                    total_sales=Sum('subtotal')
                ).order_by('-total_quantity')
                
                product_sales_data = []
                for idx, ps in enumerate(product_sales, 1):
                    # Obtener stock actual del producto
                    try:
                        product = Product.objects.get(id=ps['prod__id'])
                        stock_actual = float(product.stock or 0)
                    except:
                        stock_actual = 0
                    
                    product_sales_data.append({
                        'rank': idx,
                        'product': ps['prod__name'],
                        'quantity': int(float(ps['total_quantity'])),  # Convertir a entero
                        'price': float(ps['avg_price']),
                        'total': float(ps['total_sales']),
                        'stock': int(stock_actual)  # Stock como entero
                    })
                
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="ranking_productos_{start_date}_al_{end_date}.csv"'
                
                writer = csv.writer(response)
                writer.writerow(['REPORTE DE RANKING DE PRODUCTOS'])
                writer.writerow([f"Período: {start_date} al {end_date}"])
                writer.writerow([f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
                writer.writerow([])
                
                writer.writerow(['Ranking', 'Producto', 'Cantidad Vendida', 'Stock Actual', 'Precio Promedio', 'Total Ventas'])
                
                for product in product_sales_data:
                    writer.writerow([
                        product['rank'],
                        product['product'],
                        product['quantity'],
                        product['stock'],
                        product['price'],
                        product['total']
                    ])
                
                grand_total = sum(p['total'] for p in product_sales_data)
                total_quantity = sum(p['quantity'] for p in product_sales_data)
                
                writer.writerow([])
                writer.writerow(['Total General', '', total_quantity, '', '', grand_total])
                
                return response
            
            # CSV original para método de pago
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="ventas_{start_date}_al_{end_date}.csv"'
            
            writer = csv.writer(response)
            # Get company and user info for header
            company_name = "Todas las Empresas"
            if active_cid:
                try:
                    company = Company.objects.get(id=active_cid)
                    company_name = company.name
                except Company.DoesNotExist:
                    company_name = "Empresa Desconocida"
            
            user_name = request.user.get_full_name() or request.user.username
            
            # Add header information
            writer.writerow([f"REPORTE DE VENTAS - {company_name}"])
            writer.writerow([f"Generado por: {user_name}"])
            writer.writerow([f"Período: {start_date} al {end_date}"])
            writer.writerow([f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
            writer.writerow([])
            
            writer.writerow(['Fecha', 'Ticket/Factura', 'Cliente', 'Efectivo', 'Mercado Pago', 'Transferencias', 'Otros', 'Forma de Pago', 'Empresa'])
            
            cash_total = 0
            mp_total = 0
            transfer_total = 0
            other_total = 0
            
            for sale in sales:
                # Get invoice/ticket number
                if sale.is_invoiced:
                    ticket_number = sale.invoice_number
                else:
                    # Usar siempre el ID local para mantener consistencia
                    if hasattr(sale, 'local_sale_id') and sale.local_sale_id:
                        ticket_number = f"TK-{sale.local_sale_id:06d}"
                    else:
                        ticket_number = f"TK-{sale.id:06d}"
                
                sale_subtotal = float(sale.subtotal)  # Usar PVP puro sin IVA
                
                # Distribute amount by payment method
                cash_amount = 0
                mp_amount = 0
                transfer_amount = 0
                other_amount = 0
                
                if sale.payment_method == 'cash':
                    cash_amount = sale_subtotal
                    cash_total += cash_amount
                elif sale.payment_method == 'mp':
                    mp_amount = sale_subtotal
                    mp_total += mp_amount
                elif sale.payment_method == 'transfer':
                    transfer_amount = sale_subtotal
                    transfer_total += transfer_amount
                elif sale.payment_method in ['card', 'check']:
                    other_amount = sale_subtotal
                    other_total += other_amount
                elif sale.payment_method and '+' in sale.payment_method:
                    # Combined payments
                    payment_details = getattr(sale, 'payment_details', [])
                    if payment_details and isinstance(payment_details, list):
                        for payment in payment_details:
                            if isinstance(payment, dict):
                                method = payment.get('method', '')
                                amount = float(payment.get('amount', 0))
                                
                                if method == 'cash':
                                    cash_amount += amount
                                    cash_total += amount
                                elif method == 'mp':
                                    mp_amount += amount
                                    mp_total += amount
                                elif method == 'transfer':
                                    transfer_amount += amount
                                    transfer_total += amount
                                elif method in ['card', 'check']:
                                    other_amount += amount
                                    other_total += amount
                                else:
                                    # Método no reconocido, agregar a otros
                                    other_amount += amount
                                    other_total += amount
                    else:
                        # If no details or invalid format, put in others
                        other_amount = sale_subtotal
                        other_total += other_amount
                else:
                    # Unrecognized method, put in others
                    other_amount = sale_subtotal
                    other_total += other_amount
                
                writer.writerow([
                    sale.date_joined.strftime('%d/%m/%Y %H:%M'),
                    ticket_number,
                    sale.cli.names if sale.cli else 'Anónimo',
                    cash_amount,
                    mp_amount,
                    transfer_amount,
                    other_amount,
                    sale.get_payment_method_display(),
                    sale.company.name if sale.company else 'N/A'
                ])
            
            # Add summary
            writer.writerow([])
            writer.writerow(['RESUMEN'])
            writer.writerow(['Efectivo', cash_total])
            writer.writerow(['Mercado Pago', mp_total])
            writer.writerow(['Transferencias', transfer_total])
            writer.writerow(['Otros', other_total])
            grand_total = cash_total + mp_total + transfer_total + other_total
            writer.writerow(['Total General', grand_total])
            writer.writerow(['Cantidad de Ventas', len(sales)])
            writer.writerow(['Promedio por Venta', grand_total / len(sales) if sales else 0])
            
            # Add payment method breakdown
            writer.writerow([])
            writer.writerow(['DESGLOSE POR MÉTODO DE PAGO'])
            
            payment_choices = {
                'cash': 'Efectivo',
                'card': 'Tarjeta',
                'transfer': 'Transferencia',
                'mp': 'Mercado Pago',
                'check': 'Cheque',
                'combined': 'Combinada'
            }
            
            for method_key, method_name in payment_choices.items():
                method_sales = sales.filter(payment_method=method_key)
                if method_sales.exists():
                    method_total = method_sales.aggregate(total=Sum('total'))['total'] or 0
                    method_count = method_sales.count()
                    writer.writerow([method_name, f"${method_total:.2f}", f"({method_count} ventas)"])
            
            return response
        
        elif export_format == 'excel':
            if not OPENPYXL_AVAILABLE:
                return HttpResponse("Error: openpyxl no está instalado. Instale con: pip install openpyxl", content_type='text/plain')
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from django.http import HttpResponse
            from io import BytesIO
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="ventas_{start_date}_al_{end_date}.xlsx"'
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Reporte de Ventas'
            
            # Get company and user info for header
            company_name = "Todas las Empresas"
            if active_cid:
                try:
                    company = Company.objects.get(id=active_cid)
                    company_name = company.name
                except Company.DoesNotExist:
                    company_name = "Empresa Desconocida"
            
            user_name = request.user.get_full_name() or request.user.username
            
            # Add header information
            ws.append(['REPORTE DE VENTAS'])
            ws.merge_cells('A1:F1')
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.append([company_name])
            ws.merge_cells('A2:F2')
            ws['A2'].font = Font(bold=True, size=14)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            ws.append([''])
            ws.append(['Generado por:', user_name])
            ws.append(['Período:', f'{start_date} al {end_date}'])
            ws.append(['Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M')])
            ws.append([''])
            
            # Headers
            headers = ['Fecha', 'Ticket/Factura', 'Cliente', 'Efectivo', 'Mercado Pago', 'Transferencias', 'Otros', 'Forma de Pago', 'Empresa']
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Style header information rows
            for row_num in range(4, 8):
                for col_num in range(1, 3):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = Font(bold=True)
            
            # Style column headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col_num)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
                cell.border = thin_border
            
            # Data
            cash_total = 0
            mp_total = 0
            transfer_total = 0
            other_total = 0
            row_num = 10
            for sale in sales:
                # Get invoice/ticket number
                if sale.is_invoiced:
                    ticket_number = sale.invoice_number
                else:
                    # Usar siempre el ID local para mantener consistencia
                    if hasattr(sale, 'local_sale_id') and sale.local_sale_id:
                        ticket_number = f"TK-{sale.local_sale_id:06d}"
                    else:
                        ticket_number = f"TK-{sale.id:06d}"
                
                sale_subtotal = float(sale.subtotal)  # Usar PVP puro sin IVA
                
                # Distribute amount by payment method
                cash_amount = 0
                mp_amount = 0
                transfer_amount = 0
                other_amount = 0
                
                if sale.payment_method == 'cash':
                    cash_amount = sale_subtotal
                    cash_total += cash_amount
                elif sale.payment_method == 'mp':
                    mp_amount = sale_subtotal
                    mp_total += mp_amount
                elif sale.payment_method == 'transfer':
                    transfer_amount = sale_subtotal
                    transfer_total += transfer_amount
                elif sale.payment_method in ['card', 'check']:
                    other_amount = sale_subtotal
                    other_total += other_amount
                elif sale.payment_method and '+' in sale.payment_method:
                    # Combined payments
                    payment_details = getattr(sale, 'payment_details', [])
                    if payment_details and isinstance(payment_details, list):
                        for payment in payment_details:
                            if isinstance(payment, dict):
                                method = payment.get('method', '')
                                amount = float(payment.get('amount', 0))
                                
                                if method == 'cash':
                                    cash_amount += amount
                                    cash_total += amount
                                elif method == 'mp':
                                    mp_amount += amount
                                    mp_total += amount
                                elif method == 'transfer':
                                    transfer_amount += amount
                                    transfer_total += amount
                                elif method in ['card', 'check']:
                                    other_amount += amount
                                    other_total += amount
                                else:
                                    # Método no reconocido, agregar a otros
                                    other_amount += amount
                                    other_total += amount
                    else:
                        # If no details or invalid format, put in others
                        other_amount = sale_subtotal
                        other_total += other_amount
                else:
                    # Unrecognized method, put in others
                    other_amount = sale_subtotal
                    other_total += other_amount
                
                ws.append([
                    sale.date_joined.strftime('%d/%m/%Y %H:%M'),
                    ticket_number,
                    sale.cli.names if sale.cli else 'Anónimo',
                    cash_amount,
                    mp_amount,
                    transfer_amount,
                    other_amount,
                    sale.get_payment_method_display(),
                    sale.company.name if sale.company else 'N/A'
                ])
                
                # Style data rows
                for col_num in range(1, 10):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    if col_num in [4, 5, 6, 7]:  # Efectivo, MP, Transfer, Otros columns
                        cell.alignment = Alignment(horizontal='right')
                
                row_num += 1
            
            # Summary section
            row_num += 2
            ws.append(['RESUMEN'])
            ws.merge_cells(f'A{row_num}:I{row_num}')
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            
            row_num += 1
            ws.append(['Efectivo', cash_total])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
            
            row_num += 1
            ws.append(['Mercado Pago', mp_total])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
            
            row_num += 1
            ws.append(['Transferencias', transfer_total])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
            
            row_num += 1
            ws.append(['Otros', other_total])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
            
            row_num += 1
            grand_total = cash_total + mp_total + transfer_total + other_total
            ws.append(['Total General', grand_total])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
            
            row_num += 1
            ws.append(['Cantidad de Ventas', len(sales)])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            
            row_num += 1
            ws.append(['Promedio por Venta', grand_total / len(sales) if sales else 0])
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
            
            # Payment method breakdown
            row_num += 2
            ws.append(['DESGLOSE POR MÉTODO DE PAGO'])
            ws.merge_cells(f'A{row_num}:H{row_num}')
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            
            payment_choices = {
                'cash': 'Efectivo',
                'card': 'Tarjeta',
                'transfer': 'Transferencia',
                'mp': 'Mercado Pago',
                'check': 'Cheque',
                'combined': 'Combinada'
            }
            
            row_num += 1
            for method_key, method_name in payment_choices.items():
                method_sales = sales.filter(payment_method=method_key)
                if method_sales.exists():
                    method_total = method_sales.aggregate(total=Sum('total'))['total'] or 0
                    method_count = method_sales.count()
                    ws.append([method_name, method_total, f"({method_count} ventas)"])
                    ws.cell(row=row_num, column=1).font = Font(bold=True)
                    ws.cell(row=row_num, column=2).font = Font(bold=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
                    row_num += 1
            
            # Adjust column widths
            column_widths = [20, 15, 25, 12, 15, 15, 12, 15, 20]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            response.write(excel_file.read())
            
            return response
            
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(f"Error en exportación: {str(e)}", content_type='text/plain')


def generate_pdf_report(sales, start_date, end_date, company_id, user, report_type='daily'):
    """Generate PDF report matching the exact format from the image"""
    if not REPORTLAB_AVAILABLE:
        from django.http import HttpResponse
        return HttpResponse("Error: ReportLab no está instalado. Instale con: pip install reportlab", content_type='text/plain')
    
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from django.db.models import Sum
    from core.erp.models import Company
    
    # Obtener nombre de la empresa
    company_name = "Empresa"
    try:
        if company_id:
            company = Company.objects.get(id=company_id)
            company_name = company.name
        else:
            # Si no hay company_id, intentar obtener de la sesión o usar la primera empresa
            company = Company.objects.filter(is_active=True).first()
            if company:
                company_name = company.name
    except:
        pass
    
    # Register fonts (if available)
    try:
        pdfmetrics.registerFont(TTFont('Arial', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        font_name = 'Arial'
    except:
        font_name = 'Helvetica'
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="planilla_ventas_{start_date}_al_{end_date}.pdf"'
    
    buffer = BytesIO()
    # Use A4 with optimized margins for better space utilization
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    
    # Custom styles for the exact format
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=1,  # Center
        textColor=colors.black,
        bold=True
    )
    
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.black,
        bold=True
    )
    
    normal_style = styles['Normal']
    
    # Build story
    story = []
    
    # Header - Company name and Planilla de Ventas
    story.append(Paragraph(company_name, header_style))
    story.append(Paragraph("Planilla de Ventas", date_style))
    story.append(Spacer(1, 6))  # Reducido de 8 a 6
    
    # Date and Operator section - más compacto
    date_style_small = ParagraphStyle(
        'DateStyleSmall',
        parent=styles['Normal'],
        fontSize=10,  # Reducido de 12 a 10
        spaceAfter=6,  # Reducido espacio
        alignment=1,  # Center
        textColor=colors.black
    )
    
    # Fecha y operador en una sola línea - mostrar período según tipo de reporte
    operator_name = getattr(user, 'username', 'Operador')
    
    # Función para formatear fecha a formato latinoamericano
    def format_date_latam(date_str):
        if not date_str:
            return ''
        try:
            from datetime import datetime
            # Intentar diferentes formatos de entrada
            for fmt in ['%Y-%m-%d', '%d/%m/%Y']:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    return date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    continue
            # Si no funciona ningún formato, devolver original
            return date_str
        except:
            return date_str
    
    # Determinar el texto del período según el tipo de reporte
    if report_type == 'daily':
        period_text = f"Fecha: {format_date_latam(start_date)}"
    elif report_type == 'weekly':
        period_text = f"Período: {format_date_latam(start_date)} al {format_date_latam(end_date)}"
    elif report_type == 'monthly':
        period_text = f"Período: {format_date_latam(start_date)} al {format_date_latam(end_date)}"
    else:
        period_text = f"Período: {format_date_latam(start_date)} al {format_date_latam(end_date)}"
    
    story.append(Paragraph(f"{period_text} | Operador: {operator_name}", date_style_small))
    story.append(Spacer(1, 8))  # Reducido de 12 a 8
    
    # Función para formatear con comas
    def format_currency(amount):
        if amount == 0:
            return ''
        return f"${amount:,.2f}"
    
    # Función para formatear pagos combinados con descripción
    def format_combined_payment(amount, payment_method):
        if amount == 0:
            return ''
        # Para pagos combinados, solo mostrar el monto sin descripción
        return f"${amount:,.2f}"
    
    # Sales table with payment method columns (no Total column)
    headers = ['Producto', 'Efectivo', 'Mercado Pago', 'Transferencias', 'Otros']
    
    # Table data with payment distribution
    table_data = [headers]
    
    # Initialize totals
    cash_total = 0
    mp_total = 0
    transfer_total = 0
    other_total = 0
    
    for sale in sales:
        sale_subtotal = float(sale.subtotal)  # Usar PVP puro sin IVA
        
        # Get product names from sale details
        product_names = [det.prod.name for det in sale.detsale_set.all()]
        products_text = ", ".join(product_names) if product_names else "Sin productos"
        
        # Distribute amount by payment method
        cash_amount = 0
        mp_amount = 0
        transfer_amount = 0
        other_amount = 0
        
        if sale.payment_method == 'cash':
            cash_amount = sale_subtotal
            cash_total += cash_amount
        elif sale.payment_method == 'mp':
            mp_amount = sale_subtotal
            mp_total += mp_amount
        elif sale.payment_method == 'transfer':
            transfer_amount = sale_subtotal
            transfer_total += transfer_amount
        elif sale.payment_method in ['card', 'check']:
            other_amount = sale_subtotal
            other_total += other_amount
        elif sale.payment_method and '+' in sale.payment_method:
            # Combined payments
            payment_details = getattr(sale, 'payment_details', [])
            if payment_details and isinstance(payment_details, list):
                for payment in payment_details:
                    if isinstance(payment, dict):
                        method = payment.get('method', '')
                        amount = float(payment.get('amount', 0))
                        
                        if method == 'cash':
                            cash_amount += amount
                            cash_total += amount
                        elif method == 'mp':
                            mp_amount += amount
                            mp_total += amount
                        elif method == 'transfer':
                            transfer_amount += amount
                            transfer_total += amount
                        elif method in ['card', 'check']:
                            other_amount += amount
                            other_total += amount
                        else:
                            # Método no reconocido, agregar a otros
                            other_amount += amount
                            other_total += amount
            else:
                # If no details or invalid format, put in others
                other_amount = sale_subtotal
                other_total += other_amount
        else:
            # Unrecognized method, put in others
            other_amount = sale_subtotal
            other_total += other_amount
        
        table_data.append([
            products_text,
            format_currency(cash_amount),
            format_currency(mp_amount),
            format_currency(transfer_amount),
            format_currency(other_amount)
        ])
    
    # Add summary row
    table_data.append([
        'Resumen',
        format_currency(cash_total),
        format_currency(mp_total),
        format_currency(transfer_total),
        format_currency(other_total)
    ])
    
    # Add grand total row - igual que en el template HTML (colspan=4)
    grand_total = cash_total + mp_total + transfer_total + other_total
    table_data.append([
        'Total General',
        '',  # Efectivo vacío
        '',  # Mercado Pago vacío
        '',  # Transferencias vacío
        format_currency(grand_total)  # Total en la última columna
    ])
    
    # Create table with optimized column widths for A4 (primera columna más ancha para nombres de productos)
    sales_table = Table(table_data, colWidths=[2.0*inch, 1.1*inch, 1.1*inch, 1.1*inch, 1.1*inch])
    sales_table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOLD', (0, 0), (-1, 0), True),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Reducido de 8 a 6
        ('TOPPADDING', (0, 0), (-1, 0), 6),  # Reducido de 8 a 6
        
        # Data rows styling
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -2), font_name),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 4),  # Reducido de 6 a 4
        ('TOPPADDING', (0, 1), (-1, -2), 4),  # Reducido de 6 a 4
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # Summary row styling
        ('BACKGROUND', (0, -2), (-1, -2), colors.lightgrey),
        ('TEXTCOLOR', (0, -2), (-1, -2), colors.black),
        ('ALIGN', (0, -2), (-1, -2), 'CENTER'),
        ('FONTNAME', (0, -2), (-1, -2), font_name),
        ('FONTSIZE', (0, -2), (-1, -2), 10),
        ('BOLD', (0, -2), (-1, -2), True),
        ('BOTTOMPADDING', (0, -2), (-1, -2), 6),  # Reducido de 8 a 6
        ('TOPPADDING', (0, -2), (-1, -2), 6),  # Reducido de 8 a 6
        
        # Grand total row styling - fondo blanco, texto negro negrita
        ('BACKGROUND', (0, -1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('BOLD', (0, -1), (-1, -1), True),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ('TOPPADDING', (0, -1), (-1, -1), 8)
    ]))
    
    story.append(sales_table)
    story.append(Spacer(1, 15))  # Reducido de 20 a 15
    
    # Observations section
    obs_style = ParagraphStyle(
        'ObsStyle',
        parent=normal_style,
        fontSize=10,
        spaceAfter=10,
        leftIndent=20
    )
    
    story.append(Paragraph("Observaciones:", obs_style))
    story.append(Spacer(1, 20))
    
    # Conformity section
    conformity_style = ParagraphStyle(
        'ConformityStyle',
        parent=normal_style,
        fontSize=10,
        spaceAfter=10,
        leftIndent=20
    )
    
    story.append(Paragraph("Conformidad:", conformity_style))
    
    # Add signature line
    story.append(Spacer(1, 30))
    story.append(Paragraph("_________________________", conformity_style))
    story.append(Paragraph("Firma", conformity_style))
    
    # Add generation date
    story.append(Spacer(1, 20))
    generation_date_style = ParagraphStyle(
        'GenerationDateStyle',
        parent=normal_style,
        fontSize=9,
        spaceAfter=5,
        alignment=2,  # Right align
        textColor=colors.grey
    )
    story.append(Paragraph(f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}", generation_date_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


def generate_products_pdf_report(sales, start_date, end_date, company_id, user, report_type='daily'):
    """Generate PDF report for product ranking"""
    if not REPORTLAB_AVAILABLE:
        from django.http import HttpResponse
        return HttpResponse("Error: ReportLab no está instalado. Instale con: pip install reportlab", content_type='text/plain')
    
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from django.db.models import Sum, Avg
    from core.erp.models import Company, Product
    
    # Obtener nombre de la empresa
    company_name = "Empresa"
    try:
        if company_id:
            company = Company.objects.get(id=company_id)
            company_name = company.name
        else:
            company = Company.objects.filter(is_active=True).first()
            if company:
                company_name = company.name
    except:
        pass
    
    # Register fonts (if available)
    try:
        pdfmetrics.registerFont(TTFont('Arial', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        font_name = 'Arial'
    except:
        font_name = 'Helvetica'
    
    # Obtener datos de productos agrupados
    product_sales = DetSale.objects.filter(
        sale__in=sales
    ).values('prod__name', 'prod__id').annotate(
        total_quantity=Sum('cant'),
        avg_price=Avg('price'),
        total_sales=Sum('subtotal')
    ).order_by('-total_quantity')
    
    product_sales_data = []
    for idx, ps in enumerate(product_sales, 1):
        # Obtener stock actual del producto
        try:
            product = Product.objects.get(id=ps['prod__id'])
            stock_actual = float(product.stock or 0)
            # Debug: mostrar valores en PDF también
            print(f"PDF - Producto: {ps['prod__name']}, ID: {ps['prod__id']}, Stock: {stock_actual}")
        except Product.DoesNotExist:
            stock_actual = 0
            print(f"PDF - Producto no encontrado: ID {ps['prod__id']}")
        except Exception as e:
            stock_actual = 0
            print(f"PDF - Error obteniendo stock para {ps['prod__name']}: {e}")
        
        product_sales_data.append({
            'rank': idx,
            'product': ps['prod__name'],
            'quantity': int(float(ps['total_quantity'])),  # Convertir a entero
            'price': float(ps['avg_price']),
            'total': float(ps['total_sales']),
            'stock': int(stock_actual)  # Stock como entero
        })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ranking_productos_{start_date}_al_{end_date}.pdf"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=1,  # Center
        textColor=colors.black,
        bold=True
    )
    
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.black,
        bold=True
    )
    
    normal_style = styles['Normal']
    
    # Build story
    story = []
    
    # Header
    story.append(Paragraph(company_name, header_style))
    story.append(Paragraph("Ranking de Productos", date_style))
    story.append(Spacer(1, 6))
    
    # Date and Operator section
    date_style_small = ParagraphStyle(
        'DateStyleSmall',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        alignment=1,  # Center
        textColor=colors.black
    )
    
    operator_name = getattr(user, 'username', 'Operador')
    
    def format_date_latam(date_str):
        if not date_str:
            return ''
        try:
            from datetime import datetime
            # Intentar diferentes formatos de entrada
            for fmt in ['%Y-%m-%d', '%d/%m/%Y']:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    return date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    continue
            # Si no funciona ningún formato, devolver original
            return date_str
        except:
            return date_str
    
    if report_type == 'daily':
        period_text = f"Fecha: {format_date_latam(start_date)}"
    elif report_type == 'weekly':
        period_text = f"Período: {format_date_latam(start_date)} al {format_date_latam(end_date)}"
    elif report_type == 'monthly':
        period_text = f"Período: {format_date_latam(start_date)} al {format_date_latam(end_date)}"
    else:
        period_text = f"Período: {format_date_latam(start_date)} al {format_date_latam(end_date)}"
    
    story.append(Paragraph(f"{period_text} | Operador: {operator_name}", date_style_small))
    story.append(Spacer(1, 8))
    
    # Función para formatear con comas
    def format_currency(amount):
        if amount == 0:
            return ''
        return f"${amount:,.2f}"
    
    def format_number(amount):
        if amount == 0:
            return ''
        return f"{amount:,.2f}"
    
    def wrap_text(text, max_length=30):
        """Divide el texto en líneas si es más largo que max_length caracteres"""
        if not text or len(text) <= max_length:
            return text
        
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            if len(current_line + " " + word) <= max_length:
                if current_line:
                    current_line += " " + word
                else:
                    current_line = word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        return "\n".join(lines)
    
    # Headers para ranking de productos
    headers = ['#', 'Producto', 'Cantidad', 'Stock', 'Precio Prom.', 'Total Ventas']
    
    # Table data
    table_data = [headers]
    
    grand_total = 0
    total_quantity = 0
    
    for product in product_sales_data:
        grand_total += product['total']
        total_quantity += product['quantity']
        
        table_data.append([
            f"#{product['rank']}",
            wrap_text(product['product']),  # Aplicar wrapping de texto
            str(product['quantity']),  # Entero sin formato
            str(product['stock']),      # Entero sin formato
            format_currency(product['price']),
            format_currency(product['total'])
        ])
    
    # Add summary row
    table_data.append([
        '',
        'Total General',
        str(total_quantity),  # Entero sin formato
        '',
        '',
        format_currency(grand_total)
    ])
    
    # Create table with optimized column widths
    sales_table = Table(table_data, colWidths=[0.5*inch, 2.2*inch, 0.8*inch, 0.8*inch, 1.1*inch, 1.1*inch])
    sales_table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOLD', (0, 0), (-1, 0), True),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        # Alineación para columna de productos (índice 1) - izquierda con wrapping
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('VALIGN', (1, 1), (1, -1), 'TOP'),
        
        # Data rows styling
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -2), font_name),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 4),
        ('TOPPADDING', (0, 1), (-1, -2), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # Summary row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('BOLD', (0, -1), (-1, -1), True),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
        ('TOPPADDING', (0, -1), (-1, -1), 6)
    ]))
    
    story.append(sales_table)
    story.append(Spacer(1, 15))
    
    # Observations section
    obs_style = ParagraphStyle(
        'ObsStyle',
        parent=normal_style,
        fontSize=10,
        spaceAfter=10,
        leftIndent=20
    )
    
    story.append(Paragraph("Observaciones:", obs_style))
    story.append(Spacer(1, 20))
    
    # Conformity section
    conformity_style = ParagraphStyle(
        'ConformityStyle',
        parent=normal_style,
        fontSize=10,
        spaceAfter=10,
        leftIndent=20
    )
    
    story.append(Paragraph("Conformidad:", conformity_style))
    
    # Add signature line
    story.append(Spacer(1, 30))
    story.append(Paragraph("_________________________", conformity_style))
    story.append(Paragraph("Firma", conformity_style))
    
    # Add generation date
    story.append(Spacer(1, 20))
    generation_date_style = ParagraphStyle(
        'GenerationDateStyle',
        parent=normal_style,
        fontSize=9,
        spaceAfter=5,
        alignment=2,  # Right align
        textColor=colors.grey
    )
    story.append(Paragraph(f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}", generation_date_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response
